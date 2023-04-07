from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active
# ws = wb.create_sheet()

for row in ws.iter_rows(min_row=2):
    
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 TOP")

# 히익 성적표 오류!! 수정하겠단소리
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")
