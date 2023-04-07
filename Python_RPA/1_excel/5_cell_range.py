from openpyxl import Workbook
from random import *
wb = Workbook() 
ws = wb.active


# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개의 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column 만 가지고 오기

# print(ws["B"])
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
for cols in col_range: # 열 2개 가지고와서
    for cell in cols: # 2개를 셀로 분류
        print(cell.value)

row_title = ws[1] # 1번째 row 만 가지고 오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기 (여기선 뒤에를포함함)
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        print(cell.coordinate, end=" ") # cell의 좌표정보
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ") # 이건 튜플로 만들어줌
        # 즉 위의내용을 보면
        print(xy[0], end=" ") # A
        print(xy[1], end=" ") # 1

    print()

wb.save("sample.xlsx")