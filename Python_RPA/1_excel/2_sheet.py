from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet() # 새로운 sheet를 기본이름으로 생성 active랑다름
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경


# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 0) # 0번째 index 에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근
# new_ws.... 라는식으로 가능

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test" # 셀에 입력하는느낌쓰~
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")