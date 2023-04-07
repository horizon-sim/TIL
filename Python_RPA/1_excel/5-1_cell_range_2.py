from openpyxl import Workbook
from random import *

from openpyxl.utils.cell import column_index_from_string
wb = Workbook() 
ws = wb.active
# ws = wb.create_sheet()

ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개의 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# 전체 rows
# print(tuple(ws.rows)) # 1행씩 묶임 튜플로
# for row in tuple(ws.rows):
#     print(row[1].value) # 두번째있는 값만 출력할것이야.

# 전체 columns
# print(tuple(ws.columns)) # 1열씩 묶임 튜플로
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[1].value) # 오호 위에랑 똑같넹?

# for column in ws.iter_cols(): # 전체 columns
#     print(column[0].value)

# 1번째 줄부터 5번쨰 줄까지, 2번째 열부터 3번쨰 열까지
# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
#     print(row[0].value, row[1].value) # 범위 설정이 가능 

for col in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3):
    print(col) # 범위를 안적으면 최대임

# 행으로 값을 가져올거면 iter_rows 열로가져올거면 iter_cols

wb.save("sample.xlsx")